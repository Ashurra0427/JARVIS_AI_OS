"""
AGRO_AGENT — Excel export using openpyxl.
Requires: openpyxl>=3.1.0

Generates professional Excel workbooks for daily/weekly/monthly reporting.
All exports land in datastore/agro/exports/.
Uses asyncio.to_thread() so Excel generation never blocks the event loop.

Column layout (Jobs sheet):
  Customer Name | Location | Area Unit | Area/Qty | Time Taken | Time Unit |
  Rate/Min-Hour | Total Amount | Operator Name | Signature / Received By | Status | Notes

ENHANCED (Phase 11):
  + Revenue Analytics sheet  — service breakdown, top customers, daily trends
  + Fuel Cost Analysis sheet — fuel spend vs revenue ratio, reduction targets
  + Profit Optimization sheet — margin by service type, efficiency KPIs
"""
from __future__ import annotations

import asyncio
import logging
import os
from collections import defaultdict

logger = logging.getLogger(__name__)

OUTPUT_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "datastore", "agro", "exports"
)


def _get_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        return openpyxl, Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl>=3.1.0")


def _thin_border(Border, Side):
    return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )


def _apply_header(ws, row, cols, fill_hex="1F4E79"):
    _, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    fill = PatternFill("solid", fgColor=fill_hex)
    border = _thin_border(Border, Side)
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


JOB_FIELD_ALIASES = {
    "customer_name":  ("customer_name", "customer"),
    "location":       ("location",),
    "area_value":     ("area_value", "quantity_value"),
    "area_unit":      ("area_unit", "quantity_unit"),
    "time_taken":     ("time_taken", "time_value"),
    "time_unit":      ("time_unit",),
    # rate_per_min MUST come first: it's how per-minute (agri timer) jobs are
    # billed, which is most real-world jobs — checking rate_per_unit/rate
    # first left this column blank for all of them.
    "rate":           ("rate_per_min", "rate_per_unit", "rate"),
    "total_amount":   ("total_amount",),
    "operator_name":  ("operator_name",),
    "signature_name": ("signature_name", "received_by"),
    "status":         ("status",),
    "service":        ("service", "material"),
    "job_type":       ("job_type",),
    "scheduled_date": ("scheduled_date", "date"),
    "notes":          ("notes",),
}


def _get(j, field, default=""):
    """Resolve a logical field from a job dict, trying every known alias.

    The Flutter Job model, the jobs SQLite table, and agro_server.py handlers
    don't always agree on a field name for the same value (e.g. duration is
    `time_taken` in the DB row but `time_value` in Job.toJson()). Reading a
    single hardcoded key silently produces a blank Excel cell whenever the
    upstream name drifts. Add new aliases to JOB_FIELD_ALIASES, not here.
    """
    for key in JOB_FIELD_ALIASES.get(field, (field,)):
        v = j.get(key)
        if v not in (None, ""):
            return v
    return default


def _warn_on_schema_drift(jobs):
    """
    Defensive check run once per export: if a field is blank on EVERY job
    despite trying all known aliases, none of those aliases exist in this
    data anymore — almost always because the Flutter app or DB schema
    renamed something the exporter doesn't know about yet. This turns a
    silent blank column into a visible log line, so drift gets caught
    before a customer opens the report and finds an empty Rate/Operator
    column. When this fires, add the new field name to JOB_FIELD_ALIASES.
    """
    if not jobs:
        return
    for field, aliases in JOB_FIELD_ALIASES.items():
        if all(_get(j, field, default=None) in (None, "") for j in jobs):
            logger.warning(
                "excel_exporter: '%s' is blank on every job in this export "
                "(tried keys %s). If the app renamed this field, add the "
                "new key to JOB_FIELD_ALIASES in excel_exporter.py.",
                field, aliases,
            )


def _rate_label(j):
    is_per_min = j.get("rate_per_min") not in (None, "")
    rate = _get(j, "rate", default=None)
    if rate is None:
        return ""
    try:
        rate = float(rate)
    except (TypeError, ValueError):
        return str(rate)
    tu = (_get(j, "time_unit") or "").lower()
    if is_per_min or "min" in tu:
        return f"Rs {rate:,.2f}/min"
    if "hour" in tu:
        return f"Rs {rate:,.2f}/hr"
    unit = _get(j, "area_unit") or "unit"
    return f"Rs {rate:,.2f}/{unit}"


def _fmt_amt(v):
    if v in ("", None):
        return ""
    try:
        return f"Rs {float(v):,.0f}"
    except Exception:
        return str(v)


def _safe_float(v, default=0.0):
    try:
        return float(v or 0)
    except Exception:
        return default


JOB_HEADERS = [
    "Customer Name",
    "Location",
    "Area Unit (Katha/Bigha...)",
    "Area / Qty",
    "Time Taken",
    "Time Unit",
    "Rate / Min or Hour",
    "Total Amount (NPR)",
    "Operator Name",
    "Signature / Received By",
    "Status",
    "Service / Material",
    "Job Type",
    "Date",
    "Notes",
]
JOB_COL_WIDTHS = [22, 18, 16, 10, 10, 10, 20, 18, 18, 24, 13, 18, 12, 12, 24]
JOB_COL_LETTERS = [chr(65 + i) for i in range(len(JOB_HEADERS))]


def _job_row(j):
    return [
        _get(j, "customer_name"),
        _get(j, "location"),
        _get(j, "area_unit"),
        _get(j, "area_value"),
        _get(j, "time_taken"),
        _get(j, "time_unit"),
        _rate_label(j),
        _fmt_amt(_get(j, "total_amount", default=None)),
        _get(j, "operator_name"),
        _get(j, "signature_name"),
        str(_get(j, "status")).replace("_", " ").capitalize(),
        _get(j, "service"),
        str(_get(j, "job_type")).capitalize(),
        _get(j, "scheduled_date"),
        _get(j, "notes"),
    ]


def _write_jobs_sheet(ws2, jobs):
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    ws2.append(JOB_HEADERS)
    _apply_header(ws2, 1, JOB_COL_LETTERS)
    ws2.row_dimensions[1].height = 36
    for letter, width in zip(JOB_COL_LETTERS, JOB_COL_WIDTHS):
        ws2.column_dimensions[letter].width = width
    border = _thin_border(Border, Side)
    for j in jobs:
        ws2.append(_job_row(j))
        row_idx = ws2.max_row
        for col_num in range(1, len(JOB_HEADERS) + 1):
            ws2.cell(row=row_idx, column=col_num).border = border


# ─────────────────────────────────────────────────────────────────────────────
# Revenue Analytics sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_revenue_analytics(ws, jobs, stats, period_label):
    """Revenue breakdown by service type, top customers, and daily trend."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    GREEN  = PatternFill("solid", fgColor="00B050")
    BLUE   = PatternFill("solid", fgColor="1F4E79")
    LTBLUE = PatternFill("solid", fgColor="BDD7EE")
    border = _thin_border(Border, Side)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # ── Title ─────────────────────────────────────────────────────────────────
    ws["A1"] = f"REVENUE ANALYTICS — {period_label}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = "JARVIS AGRO · Nawal Parasi, Nepal"
    ws["A2"].font = Font(size=10, italic=True, color="7F7F7F")
    ws.append([])

    # ── KPI Summary ───────────────────────────────────────────────────────────
    total_revenue = _safe_float(stats.get("revenue"))
    total_jobs    = int(stats.get("total_jobs", 0) or 0)
    completed     = int(stats.get("completed_jobs", 0) or 0)
    avg_revenue   = (total_revenue / completed) if completed else 0

    ws["A4"] = "📊 KEY REVENUE METRICS"
    ws["A4"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[4].height = 22

    kpis = [
        ("Total Revenue (NPR)",     _fmt_amt(total_revenue)),
        ("Completed Jobs",          completed),
        ("Total Jobs Scheduled",    total_jobs),
        ("Avg Revenue / Job (NPR)", _fmt_amt(avg_revenue)),
        ("Collection Rate",         f"{(completed/total_jobs*100) if total_jobs else 0:.1f}%"),
    ]
    ws.append(["Metric", "Value"])
    _apply_header(ws, ws.max_row, ["A", "B"])
    for label, value in kpis:
        ws.append([label, value])
        r = ws.max_row
        ws[f"A{r}"].border = border
        ws[f"B{r}"].border = border
        ws[f"B{r}"].alignment = Alignment(horizontal="right")

    ws.append([])

    # ── Revenue by Service Type ────────────────────────────────────────────────
    service_rev: dict[str, float] = defaultdict(float)
    service_cnt: dict[str, int]   = defaultdict(int)
    for j in jobs:
        if j.get("status") == "completed":
            svc = str(_get(j, "service", default="Other") or "Other").strip()
            amt = _safe_float(j.get("total_amount"))
            service_rev[svc] += amt
            service_cnt[svc] += 1

    ws_row = ws.max_row + 1
    ws[f"A{ws_row}"] = "📋 REVENUE BY SERVICE TYPE"
    ws[f"A{ws_row}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[ws_row].height = 22
    ws_row += 1
    headers = ["Service", "Jobs Done", "Revenue (NPR)", "Avg/Job (NPR)", "% of Total"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=ws_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = BLUE
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    ws_row += 1

    sorted_services = sorted(service_rev.items(), key=lambda x: x[1], reverse=True)
    for svc, rev in sorted_services:
        cnt = service_cnt[svc]
        avg = rev / cnt if cnt else 0
        pct = (rev / total_revenue * 100) if total_revenue else 0
        row_data = [svc, cnt, _fmt_amt(rev), _fmt_amt(avg), f"{pct:.1f}%"]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_idx, value=val)
            cell.border = border
            if col_idx in (3, 4):
                cell.alignment = Alignment(horizontal="right")
        ws_row += 1

    ws_row += 1

    # ── Top Customers ─────────────────────────────────────────────────────────
    cust_rev: dict[str, float] = defaultdict(float)
    cust_cnt: dict[str, int]   = defaultdict(int)
    for j in jobs:
        if j.get("status") == "completed":
            name = str(_get(j, "customer_name", default="Unknown") or "Unknown").strip()
            cust_rev[name] += _safe_float(j.get("total_amount"))
            cust_cnt[name] += 1

    ws[f"A{ws_row}"] = "🏆 TOP CUSTOMERS"
    ws[f"A{ws_row}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[ws_row].height = 22
    ws_row += 1
    headers2 = ["Customer", "Jobs Done", "Total Revenue (NPR)", "Avg/Job (NPR)"]
    for col_idx, h in enumerate(headers2, 1):
        cell = ws.cell(row=ws_row, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = BLUE
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    ws_row += 1

    top_customers = sorted(cust_rev.items(), key=lambda x: x[1], reverse=True)[:10]
    for rank, (cust, rev) in enumerate(top_customers, 1):
        cnt = cust_cnt[cust]
        avg = rev / cnt if cnt else 0
        row_data = [f"#{rank}  {cust}", cnt, _fmt_amt(rev), _fmt_amt(avg)]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=ws_row, column=col_idx, value=val)
            cell.border = border
            if col_idx in (3, 4):
                cell.alignment = Alignment(horizontal="right")
            if rank == 1:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
                cell.font = Font(bold=True)
        ws_row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Fuel Cost Analysis & Reduction Targets sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_fuel_analysis(ws, stats, jobs, fuel_logs=None):
    """Fuel spend vs revenue ratio, per-job fuel cost, and reduction targets."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    RED    = PatternFill("solid", fgColor="C00000")
    ORANGE = PatternFill("solid", fgColor="FF8C00")
    GREEN  = PatternFill("solid", fgColor="00B050")
    BLUE   = PatternFill("solid", fgColor="1F4E79")
    border = _thin_border(Border, Side)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # ── Title ─────────────────────────────────────────────────────────────────
    ws["A1"] = "⛽ FUEL COST ANALYSIS & REDUCTION TARGETS"
    ws["A1"].font = Font(bold=True, size=14, color="C00000")
    ws["A2"] = "JARVIS AGRO · Nawal Parasi, Nepal"
    ws["A2"].font = Font(size=10, italic=True, color="7F7F7F")
    ws.append([])

    revenue    = _safe_float(stats.get("revenue"))
    fuel_cost  = _safe_float(stats.get("fuel_cost"))
    other_exp  = _safe_float(stats.get("other_expenses"))
    profit     = _safe_float(stats.get("profit"))
    total_jobs = int(stats.get("total_jobs", 1) or 1)
    completed  = int(stats.get("completed_jobs", 1) or 1) or 1

    fuel_ratio = (fuel_cost / revenue * 100) if revenue else 0
    cost_per_job = fuel_cost / completed

    # ── KPIs ──────────────────────────────────────────────────────────────────
    ws["A4"] = "⛽ FUEL COST KPIs"
    ws["A4"].font = Font(bold=True, size=11, color="C00000")
    ws.row_dimensions[4].height = 22

    ws.append(["Metric", "Current", "Target", "Status"])
    _apply_header(ws, ws.max_row, ["A", "B", "C", "D"])

    def _kpi_row(label, current_val, target_val, current_fmt, target_fmt, lower_is_better=True):
        r = ws.max_row + 1
        ws.cell(r, 1, label).border = border
        ws.cell(r, 2, current_fmt).border = border
        ws.cell(r, 3, target_fmt).border = border
        if lower_is_better:
            ok = current_val <= target_val
        else:
            ok = current_val >= target_val
        status_cell = ws.cell(r, 4, "✅ On Track" if ok else "⚠️ Over Target")
        status_cell.border = border
        status_cell.fill = GREEN if ok else ORANGE
        status_cell.font = Font(bold=True, color="FFFFFF")
        ws.cell(r, 2).alignment = Alignment(horizontal="right")
        ws.cell(r, 3).alignment = Alignment(horizontal="right")

    _kpi_row("Fuel Cost (NPR)",        fuel_cost,   revenue * 0.15,  _fmt_amt(fuel_cost),    _fmt_amt(revenue * 0.15))
    _kpi_row("Fuel % of Revenue",      fuel_ratio,  15.0,            f"{fuel_ratio:.1f}%",   "15.0%")
    _kpi_row("Fuel Cost per Job (NPR)", cost_per_job, 500,           _fmt_amt(cost_per_job), _fmt_amt(500))
    _kpi_row("Net Profit (NPR)",        profit,      revenue * 0.60, _fmt_amt(profit),       _fmt_amt(revenue * 0.60), lower_is_better=False)

    ws.append([])

    # ── Reduction Recommendations ──────────────────────────────────────────────
    r = ws.max_row + 1
    ws[f"A{r}"] = "🔧 FUEL REDUCTION RECOMMENDATIONS"
    ws[f"A{r}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[r].height = 22
    r += 1

    recs = []
    if fuel_ratio > 20:
        recs.append(("🔴 HIGH PRIORITY", "Fuel exceeds 20% of revenue — review route planning immediately",
                      f"Save Rs {fuel_cost * 0.20:,.0f}/period by cutting 20%"))
    elif fuel_ratio > 15:
        recs.append(("🟡 MEDIUM", "Fuel is 15–20% of revenue — optimize job clustering by area",
                      f"Save Rs {fuel_cost * 0.10:,.0f}/period by cutting 10%"))
    else:
        recs.append(("🟢 GOOD", "Fuel cost is within target (<15% of revenue)", "Maintain current practices"))

    recs += [
        ("💡 Route Optimization",  "Cluster same-day jobs by location (ward/tole) to reduce travel",
         "Est. 10–15% fuel saving"),
        ("💡 Full Loads Only",     "For transport jobs: never run Tali/Trip below 80% capacity",
         "Est. 8–12% fuel saving"),
        ("💡 Idle Time Reduction", "Operators: switch off engine during loading/waiting (>5 min idle)",
         "Est. 3–5% fuel saving"),
        ("💡 Maintenance",         "Regular tractor service every 250 hrs prevents excess fuel burn",
         "Est. 5–8% fuel saving"),
        ("💡 Pump Comparison",     "Compare prices at local petrol pumps monthly — log in Fuel section",
         "Est. 2–4% fuel saving"),
    ]

    ws.cell(r, 1, "Priority").border = border
    ws.cell(r, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 1).fill = BLUE
    ws.cell(r, 2, "Recommendation").border = border
    ws.cell(r, 2).font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 2).fill = BLUE
    ws.cell(r, 3, "Potential Saving").border = border
    ws.cell(r, 3).font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 3).fill = BLUE
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 28
    r += 1

    for priority, rec, saving in recs:
        ws.cell(r, 1, priority).border = border
        ws.cell(r, 2, rec).border = border
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws.cell(r, 3, saving).border = border
        ws.row_dimensions[r].height = 30
        r += 1

    ws.append([])

    # ── Cost Structure Breakdown ───────────────────────────────────────────────
    r = ws.max_row + 1
    ws[f"A{r}"] = "📊 COST STRUCTURE"
    ws[f"A{r}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[r].height = 22
    r += 1

    cost_rows = [
        ("Revenue",         revenue,   100.0),
        ("Fuel Cost",       fuel_cost, (fuel_cost / revenue * 100) if revenue else 0),
        ("Other Expenses",  other_exp, (other_exp / revenue * 100) if revenue else 0),
        ("Net Profit",      profit,    (profit / revenue * 100) if revenue else 0),
    ]
    ws.cell(r, 1, "Item").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 1).fill = BLUE
    ws.cell(r, 1).border = border
    ws.cell(r, 2, "Amount (NPR)").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 2).fill = BLUE
    ws.cell(r, 2).border = border
    ws.cell(r, 3, "% of Revenue").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 3).fill = BLUE
    ws.cell(r, 3).border = border
    r += 1

    for label, amount, pct in cost_rows:
        ws.cell(r, 1, label).border = border
        amt_cell = ws.cell(r, 2, _fmt_amt(amount))
        amt_cell.border = border
        amt_cell.alignment = Alignment(horizontal="right")
        pct_cell = ws.cell(r, 3, f"{pct:.1f}%")
        pct_cell.border = border
        pct_cell.alignment = Alignment(horizontal="right")
        if label == "Net Profit":
            color = "00B050" if amount >= 0 else "C00000"
            ws.cell(r, 1).font = Font(bold=True)
            amt_cell.font = Font(bold=True, color=color)
            pct_cell.font = Font(bold=True, color=color)
        r += 1


# ─────────────────────────────────────────────────────────────────────────────
# Profit Optimization sheet
# ─────────────────────────────────────────────────────────────────────────────

def _write_profit_optimization(ws, jobs, stats):
    """Margin by service type, top/bottom performers, efficiency KPIs."""
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    BLUE  = PatternFill("solid", fgColor="1F4E79")
    GREEN = PatternFill("solid", fgColor="00B050")
    RED   = PatternFill("solid", fgColor="C00000")
    border = _thin_border(Border, Side)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    ws["A1"] = "📈 PROFIT OPTIMIZATION DASHBOARD"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = "JARVIS AGRO · Nawal Parasi, Nepal"
    ws["A2"].font = Font(size=10, italic=True, color="7F7F7F")
    ws.append([])

    revenue   = _safe_float(stats.get("revenue"))
    fuel_cost = _safe_float(stats.get("fuel_cost"))
    profit    = _safe_float(stats.get("profit"))

    # ── Margin by Service ──────────────────────────────────────────────────────
    ws["A4"] = "🏭 REVENUE BY SERVICE (High to Low)"
    ws["A4"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[4].height = 22

    service_data: dict[str, dict] = defaultdict(lambda: {"jobs": 0, "revenue": 0.0})
    for j in jobs:
        if j.get("status") == "completed":
            svc = str(_get(j, "service", default="Other") or "Other").strip()
            service_data[svc]["jobs"] += 1
            service_data[svc]["revenue"] += _safe_float(j.get("total_amount"))

    ws.append(["Service", "Jobs", "Revenue (NPR)", "Avg/Job (NPR)", "Revenue Rank"])
    _apply_header(ws, ws.max_row, ["A", "B", "C", "D", "E"])

    sorted_svcs = sorted(service_data.items(), key=lambda x: x[1]["revenue"], reverse=True)
    for rank, (svc, data) in enumerate(sorted_svcs, 1):
        avg = data["revenue"] / data["jobs"] if data["jobs"] else 0
        r = ws.max_row + 1
        ws.cell(r, 1, svc).border = border
        ws.cell(r, 2, data["jobs"]).border = border
        ws.cell(r, 3, _fmt_amt(data["revenue"])).border = border
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
        ws.cell(r, 4, _fmt_amt(avg)).border = border
        ws.cell(r, 4).alignment = Alignment(horizontal="right")
        rank_cell = ws.cell(r, 5, f"#{rank}")
        rank_cell.border = border
        rank_cell.alignment = Alignment(horizontal="center")
        if rank == 1:
            for c in range(1, 6):
                ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF2CC")
                ws.cell(r, c).font = Font(bold=True)

    ws.append([])

    # ── Efficiency KPIs ────────────────────────────────────────────────────────
    r = ws.max_row + 1
    ws[f"A{r}"] = "⚡ EFFICIENCY KPIs"
    ws[f"A{r}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[r].height = 22
    r += 1

    completed = int(stats.get("completed_jobs", 0) or 0)
    total     = int(stats.get("total_jobs", 0) or 1) or 1
    completion_rate = completed / total * 100

    kpis = [
        ("Job Completion Rate",       f"{completion_rate:.1f}%",
         "✅ Excellent" if completion_rate >= 90 else "⚠️ Needs Attention"),
        ("Revenue per Completed Job", _fmt_amt(revenue / completed if completed else 0), ""),
        ("Fuel Cost as % Revenue",    f"{(fuel_cost/revenue*100) if revenue else 0:.1f}%",
         "✅ Good (<15%)" if (fuel_cost / revenue * 100 if revenue else 100) < 15 else "⚠️ Reduce Fuel"),
        ("Net Profit Margin",         f"{(profit/revenue*100) if revenue else 0:.1f}%",
         "✅ Healthy (>50%)" if (profit / revenue * 100 if revenue else 0) >= 50 else "⚠️ Below Target"),
    ]

    ws.cell(r, 1, "KPI").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 1).fill = BLUE
    ws.cell(r, 1).border = border
    ws.cell(r, 2, "Value").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 2).fill = BLUE
    ws.cell(r, 2).border = border
    ws.cell(r, 3, "Assessment").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 3).fill = BLUE
    ws.cell(r, 3).border = border
    ws.column_dimensions["C"].width = 30
    r += 1

    for kpi_name, value, assessment in kpis:
        ws.cell(r, 1, kpi_name).border = border
        v_cell = ws.cell(r, 2, value)
        v_cell.border = border
        v_cell.alignment = Alignment(horizontal="right")
        a_cell = ws.cell(r, 3, assessment)
        a_cell.border = border
        if "✅" in assessment:
            a_cell.font = Font(color="00B050", bold=True)
        elif "⚠️" in assessment:
            a_cell.font = Font(color="FF8C00", bold=True)
        r += 1

    ws.append([])

    # ── Action Plan ───────────────────────────────────────────────────────────
    r = ws.max_row + 1
    ws[f"A{r}"] = "🎯 PROFIT IMPROVEMENT ACTION PLAN"
    ws[f"A{r}"].font = Font(bold=True, size=11, color="1F4E79")
    ws.row_dimensions[r].height = 22
    r += 1

    actions = []
    if sorted_svcs:
        top_svc = sorted_svcs[0][0]
        actions.append(("Prioritize High-Revenue Services",
                         f"Focus on '{top_svc}' jobs first — highest revenue generator",
                         "HIGH"))
    if completion_rate < 90:
        actions.append(("Improve Completion Rate",
                         f"Current {completion_rate:.0f}% — confirm jobs 1 day ahead to avoid cancellations",
                         "HIGH"))
    actions += [
        ("Advance Payment Policy",   "Collect 30–50% advance on booking to reduce bad debt risk", "MEDIUM"),
        ("Seasonal Planning",         "Plan tractor maintenance in off-season (monsoon lull) to maximize uptime", "MEDIUM"),
        ("Record All Transactions",   "Log every fuel purchase and expense immediately for accurate profit tracking", "HIGH"),
        ("Monthly Report Review",     "Review monthly Excel report every 1st of month to spot trends early", "LOW"),
    ]

    ws.cell(r, 1, "Action").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 1).fill = BLUE
    ws.cell(r, 1).border = border
    ws.cell(r, 2, "Details").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 2).fill = BLUE
    ws.cell(r, 2).border = border
    ws.column_dimensions["B"].width = 55
    ws.cell(r, 3, "Priority").font = Font(bold=True, color="FFFFFF")
    ws.cell(r, 3).fill = BLUE
    ws.cell(r, 3).border = border
    ws.column_dimensions["C"].width = 12
    r += 1

    for action_name, detail, priority in actions:
        ws.cell(r, 1, action_name).border = border
        d_cell = ws.cell(r, 2, detail)
        d_cell.border = border
        d_cell.alignment = Alignment(wrap_text=True)
        p_cell = ws.cell(r, 3, priority)
        p_cell.border = border
        p_cell.alignment = Alignment(horizontal="center")
        if priority == "HIGH":
            p_cell.fill = RED
            p_cell.font = Font(bold=True, color="FFFFFF")
        elif priority == "MEDIUM":
            p_cell.fill = PatternFill("solid", fgColor="FF8C00")
            p_cell.font = Font(bold=True, color="FFFFFF")
        ws.row_dimensions[r].height = 28
        r += 1


# ── Daily report ──────────────────────────────────────────────────────────────

async def generate_daily_report(stats, jobs, date_str):
    return await asyncio.to_thread(_daily_sync, stats, jobs, date_str)


def _daily_sync(stats, jobs, date_str):
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    _warn_on_schema_drift(jobs)
    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Daily Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws["A1"] = "JARVIS AGRO — Daily Report"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = f"Date: {date_str}  |  Nawal Parasi, Nepal"
    ws["A2"].font = Font(size=10, italic=True)
    ws.append([])
    ws.append(["Metric", "Value"])
    _apply_header(ws, ws.max_row, ["A", "B"])
    summary_rows = [
        ("Total Jobs Scheduled", stats.get("total_jobs", 0)),
        ("Completed Jobs",       stats.get("completed_jobs", 0)),
        ("Pending / Confirmed",  stats.get("pending_jobs", 0)),
        ("Revenue (NPR)",        _fmt_amt(stats.get("revenue", 0))),
        ("Fuel Cost (NPR)",      _fmt_amt(stats.get("fuel_cost", 0))),
        ("Other Expenses (NPR)", _fmt_amt(stats.get("other_expenses", 0))),
        ("Total Expenses (NPR)", _fmt_amt(stats.get("total_expenses", 0))),
        ("Net Profit (NPR)",     _fmt_amt(stats.get("profit", 0))),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
    profit_row = ws.max_row
    profit = stats.get("profit", 0)
    ws[f"B{profit_row}"].font = Font(bold=True, color="00B050" if (profit or 0) >= 0 else "FF0000")

    # Sheet 2: Jobs (full column layout)
    ws2 = wb.create_sheet("Jobs")
    _write_jobs_sheet(ws2, jobs)

    # Sheet 3: Revenue Analytics (NEW)
    ws_rev = wb.create_sheet("Revenue Analytics")
    _write_revenue_analytics(ws_rev, jobs, stats, f"Daily — {date_str}")

    # Sheet 4: Fuel Cost Analysis (NEW)
    ws_fuel = wb.create_sheet("Fuel Cost Analysis")
    _write_fuel_analysis(ws_fuel, stats, jobs)

    # Sheet 5: Profit Optimization (NEW)
    ws_profit = wb.create_sheet("Profit Optimization")
    _write_profit_optimization(ws_profit, jobs, stats)

    # Sheet 6: Receipt Stubs (printable, one per job)
    ws3 = wb.create_sheet("Receipt Stubs")
    ws3["A1"] = "JARVIS AGRO — Daily Receipts"
    ws3["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws3["A2"] = f"Date: {date_str}"
    ws3["A2"].font = Font(size=10, italic=True)
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 34
    stub_row = 4
    for idx, j in enumerate(jobs, 1):
        service    = _get(j, "service")
        area_val   = _get(j, "area_value")
        area_unit  = _get(j, "area_unit")
        time_taken = _get(j, "time_taken")
        time_unit  = _get(j, "time_unit")
        total_amt  = _get(j, "total_amount", default=0) or 0
        lines = [
            (f"Job #{idx}  —  {service}", ""),
            ("Customer:", _get(j, "customer_name")),
            ("Location:", _get(j, "location")),
            ("Operator:", _get(j, "operator_name")),
        ]
        if time_taken:
            lines.append(("Time Taken:", f"{time_taken} {time_unit}"))
        if area_val:
            lines.append(("Area / Qty:", f"{area_val} {area_unit}"))
        lines += [
            ("Rate:", _rate_label(j)),
            ("Total Amount:", f"Rs {float(total_amt):,.0f}"),
            ("Signature:", "___________________________"),
            ("", ""),
        ]
        for label, value in lines:
            ws3.cell(row=stub_row, column=1, value=label).font = Font(bold=bool(label and ":" in label or idx > 0 and label.startswith("Job")))
            ws3.cell(row=stub_row, column=2, value=value)
            stub_row += 1
        ws3.cell(row=stub_row, column=1, value="─" * 50)
        stub_row += 2

    filename = f"daily_report_{date_str}.xlsx"
    path = os.path.abspath(os.path.join(OUTPUT_DIR, filename))
    wb.save(path)
    return path


# ── Monthly report ────────────────────────────────────────────────────────────

async def generate_monthly_report(month, daily_stats, all_jobs, all_expenses):
    return await asyncio.to_thread(_monthly_sync, month, daily_stats, all_jobs, all_expenses)


def _monthly_sync(month, daily_stats, all_jobs, all_expenses):
    openpyxl, Font, PatternFill, Alignment, Border, Side = _get_openpyxl()
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    _warn_on_schema_drift(all_jobs)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Monthly Summary"
    ws["A1"] = f"JARVIS AGRO — Monthly Report: {month}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = "Nawal Parasi, Nepal"
    ws["A2"].font = Font(size=10, italic=True)
    ws.append([])
    headers = ["Date", "Total Jobs", "Completed", "Revenue (NPR)", "Fuel (NPR)", "Expenses (NPR)", "Profit (NPR)"]
    ws.append(headers)
    _apply_header(ws, 4, [chr(65 + i) for i in range(len(headers))])
    for s in daily_stats:
        ws.append([
            s.get("date"), s.get("total_jobs", 0), s.get("completed_jobs", 0),
            s.get("revenue", 0), s.get("fuel_cost", 0), s.get("other_expenses", 0), s.get("profit", 0),
        ])
    last_data_row = 4 + len(daily_stats)
    total_row = last_data_row + 1
    ws.append(["TOTAL"] + [f"=SUM({c}5:{c}{last_data_row})" for c in "BCDEFG"])
    ws[f"A{total_row}"].font = Font(bold=True)
    for col in "BCDEFG":
        ws[f"{col}{total_row}"].font = Font(bold=True)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(str(c.value or "")) for c in col), default=10) + 4, 30
        )

    # ── Aggregate monthly stats for analytics sheets ──────────────────────────
    monthly_agg = {
        "revenue":        sum(_safe_float(s.get("revenue")) for s in daily_stats),
        "fuel_cost":      sum(_safe_float(s.get("fuel_cost")) for s in daily_stats),
        "other_expenses": sum(_safe_float(s.get("other_expenses")) for s in daily_stats),
        "total_jobs":     sum(int(s.get("total_jobs", 0) or 0) for s in daily_stats),
        "completed_jobs": sum(int(s.get("completed_jobs", 0) or 0) for s in daily_stats),
        "pending_jobs":   sum(int(s.get("pending_jobs", 0) or 0) for s in daily_stats),
    }
    monthly_agg["total_expenses"] = monthly_agg["fuel_cost"] + monthly_agg["other_expenses"]
    monthly_agg["profit"] = monthly_agg["revenue"] - monthly_agg["total_expenses"]

    ws2 = wb.create_sheet("All Jobs")
    _write_jobs_sheet(ws2, all_jobs)

    # Revenue Analytics (NEW)
    ws_rev = wb.create_sheet("Revenue Analytics")
    _write_revenue_analytics(ws_rev, all_jobs, monthly_agg, f"Monthly — {month}")

    # Fuel Cost Analysis (NEW)
    ws_fuel = wb.create_sheet("Fuel Cost Analysis")
    _write_fuel_analysis(ws_fuel, monthly_agg, all_jobs, all_expenses)

    # Profit Optimization (NEW)
    ws_profit = wb.create_sheet("Profit Optimization")
    _write_profit_optimization(ws_profit, all_jobs, monthly_agg)

    if all_expenses:
        ws3 = wb.create_sheet("Expenses")
        exp_headers = ["ID", "Category", "Amount (NPR)", "Job ID", "Description", "Date"]
        ws3.append(exp_headers)
        _apply_header(ws3, 1, ["A", "B", "C", "D", "E", "F"])
        for e in all_expenses:
            ws3.append([
                e.get("id"), e.get("category", ""), e.get("amount", 0),
                e.get("job_id", ""), e.get("description", ""), e.get("logged_at", ""),
            ])

    filename = f"monthly_report_{month}.xlsx"
    path = os.path.abspath(os.path.join(OUTPUT_DIR, filename))
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Class wrapper — kept for backwards/forwards compatibility.
#
# agro_server.py imports this module as:
#     from agents.agro.excel_exporter import AgroExcelExporter as xl
#     await xl.generate_daily_report(...)
#     await xl.generate_monthly_report(...)
#
# The actual implementation above is a set of plain module-level functions
# (no class), which caused:
#   ImportError: cannot import name 'AgroExcelExporter' from
#   'agents.agro.excel_exporter'
#
# Rather than rewrite every call site in agro_server.py, this thin wrapper
# exposes the existing module functions as staticmethods so the expected
# `AgroExcelExporter` symbol exists and behaves identically.
# ---------------------------------------------------------------------------

class AgroExcelExporter:
    """Static-method facade over the module-level report generators."""

    generate_daily_report = staticmethod(generate_daily_report)
    generate_monthly_report = staticmethod(generate_monthly_report)
